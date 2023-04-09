from openpyxl import load_workbook, Workbook

#Lista de arquivos para ser consolidada
caminho = 'AUTOMACAO_PYTHON_EXCEL\\ROBO_UNIFICANDO_INFORMACOES\\'
lista_arquivos = ['CustosAutom', 'PopulacaoPOA', 'SuperMercado']


wb = Workbook()

arquivo_final = 'AUTOMACAO_PYTHON_EXCEL\\ROBO_UNIFICANDO_INFORMACOES\\resultado.xlsx'


for nome_arquivo in lista_arquivos:

    arquivo = load_workbook(filename='%s\\%s.xlsx' % (caminho, nome_arquivo))
    sheet = arquivo[nome_arquivo]
    max_linhas = sheet.max_row
    max_colunhas = sheet.max_column
    
    ws = wb.create_sheet(title=nome_arquivo)
    
    
    #Passar os dados de um arquivo para o outros
    for i in range(1, max_linhas + 1):
        for j in range(1, max_colunhas + 1):
            cel = sheet.cell(row=i, column=j)
            ws.cell(row=i, column= j).value = cel.value
wb.remove(wb['Sheet'])            
wb.save(arquivo_final)