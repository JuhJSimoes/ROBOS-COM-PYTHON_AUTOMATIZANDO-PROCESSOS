from openpyxl import Workbook, load_workbook
from openpyxl.chart import BarChart, Reference, series

dict_anos = {}

arquivo1 = load_workbook(filename='AUTOMACAO_PYTHON_EXCEL\\ROBO_GRAFICO_2D\\gastos.xlsx')
ws1 = arquivo1['gastos']
max_linhas = ws1.max_row

for i in range(2, max_linhas + 1):
    dict_anos[ws1['A%s' % i].value] = {'gastos' : ws1['B%s' %i].value, 'receita' : 0}
    
print(dict_anos)