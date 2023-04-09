from openpyxl import load_workbook

wb = load_workbook('AUTOMACAO_PYTHON_EXCEL\\TRABALHANDO_COM_FORMULAS\\nomes.xlsx')

planilha = wb['Planilha1']


soma_idade = 0
for i in range(2,5):
    print("%s tem %s anos." % (planilha['A%s' % i].value, planilha['B%s' % i].value))
    soma_idade += int(planilha['B%s' %i].value)


#Opção de soma utilizando formula do Excel  
#planilha['B5'] = '=SUM(B2:B4)'

planilha['B5'] = soma_idade

#Mesclando Celula e removendo
planilha['A7'] = 'ALUNOS'
planilha.merge_cells('A7:B7')
planilha.unmerge_cells('A7:B7')

wb.save(filename='AUTOMACAO_PYTHON_EXCEL\\TRABALHANDO_COM_FORMULAS\\nomes.xlsx')